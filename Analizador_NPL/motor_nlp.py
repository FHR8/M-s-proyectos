# -*- coding: utf-8 -*-.
import os
import re
import logging
#Counter -> Importamos el contador.
from collections import Counter
#Pandas -> Para la manipulación de datos estadísticos.
import pandas as pd
# Docx -> Para trabajar con documentos Word
from docx import Document
# Pdfplumber -> Para trabajar con documentos PDF
import pdfplumber
# Spacy -> Librería principal encargada de trabajar con lenguaje natural
import spacy
# STOP_WORDS -> Permite sólo obtener las palabras que tienen un significado propio (quitando preposiciones, conjunciones...)
from spacy.lang.es.stop_words import STOP_WORDS
# openpyxl -> Permite el trabajo con el excel (datos, colores, gráficos)
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.chart import PieChart, BarChart, Reference
from openpyxl.formatting.rule import ColorScaleRule
#Matplotlib -> Para el uso de gráficos en el programa
import matplotlib
matplotlib.use('Agg') 
import matplotlib.pyplot as plt
#Para la creación de la imagen de nube de palabras
from wordcloud import WordCloud

#Configuración del logging para silenciarlo
#Si ocurre un error en el gestor de archivos, nos dará la información de interés con fecha y hora.
logging.basicConfig(filename='app_nlp.log', level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s', encoding='utf-8')

class GestorArchivos:
    #Basicamente, se encarga de comprobar la lectura de los archivos a analizar. Si son 
    #Word o PDF los aceptará para continuar con el proceso, y si no nos dará un error.
    @staticmethod
    def leer_documento(ruta_archivo):
        try:
            if ruta_archivo.endswith('.docx'): return GestorArchivos._leer_docx(ruta_archivo)
            elif ruta_archivo.endswith('.pdf'): return GestorArchivos._leer_pdf(ruta_archivo)
            else: raise ValueError(f"Formato no soportado: {ruta_archivo}")
        except Exception as e:
            logging.error(f"Fallo al leer el documento {ruta_archivo}: {e}"); raise 
    #Método concreto que se utiliza para leer en caso de documento word
    @staticmethod
    def _leer_docx(archivo):
        return '\n'.join([p.text for p in Document(archivo).paragraphs])
    #Método concreto que se utiliza para leer en caso de PDF (llevan procesos diferentes)
    @staticmethod
    def _leer_pdf(archivo):
        texto = ''
        with pdfplumber.open(archivo) as pdf:
            for pagina in pdf.pages: texto += (pagina.extract_text() or "") + '\n'
        return texto
    #Para la contabilización de palabras y el análisis, requerimos de texto plano, por lo que 
    #Convertimos nuestros PDF o Word en Txt, a partir del cual se trabaja.
    @staticmethod
    def guardar_texto_plano(texto, ruta_destino):
        try:
            with open(ruta_destino, 'w', encoding='utf-8') as f: f.write(texto)
        except Exception as e: logging.error(f"Error al guardar texto temporal: {e}")

#Cargamos la IA ('es_core_news_sm') que es la que se encarga del tratamiento de las palabras
#Normalizamos algunas palabras para no contabilizar los plurales por separado de los singulares.
class ProcesadorNLP:
    def __init__(self):
        try:
            self.nlp = spacy.load("es_core_news_sm")
            self.normalizacion = {"derechos": "derecho", "leyes": "ley", "personas": "persona", "niños": "niño"}
        except Exception as e: logging.critical(f"Error fatal al cargar Spacy: {e}"); raise

#Convertimos mayusculas a minúsculas y eliminamos caracteres extraños
#Fulminamos las palabras sin significado semántico.
    def limpiar_texto(self, texto):
        texto = re.sub(r'[^a-záéíóúüñ\s]', '', texto.lower())
        return " ".join([p for p in texto.split() if p not in STOP_WORDS])

#Tilizamos el texto limpio (txt), para contabilizar sustantivos, adejetivos...
#Usamos el texto original para buscar entidades ("Banco de España")
    def clasificar_palabras(self, texto_original, texto_limpio):
        doc_limpio = self.nlp(texto_limpio)
        cat = {"Sustantivos": Counter(), "Verbos": Counter(), "Adjetivos": Counter(), 
               "Adverbios": Counter(), "Entidades (NER)": Counter(), "Total": Counter()}
        
        for t in doc_limpio:
            if not t.is_alpha: continue
            lem = self.normalizacion.get(t.lemma_, t.lemma_)
            p = t.pos_ 
            if p == "NOUN": cat["Sustantivos"][lem] += 1
            elif p == "VERB": cat["Verbos"][lem.split(" ")[0] if " " in lem else lem] += 1
            elif p == "ADJ": cat["Adjetivos"][lem] += 1
            elif p == "ADV": cat["Adverbios"][lem] += 1
            cat["Total"][lem] += 1

#Con entidades nor referimos a Personas, localizaciones y organizaciones
#Como las recogemos del original, tenemos que hacer una limpieza concreta
        doc_original = self.nlp(texto_original)
        for ent in doc_original.ents:
            if ent.label_ in ['PER', 'LOC', 'ORG']:
                entidad_limpia = ent.text.replace('\n', ' ').strip().title()
                if len(entidad_limpia) > 2: cat["Entidades (NER)"][entidad_limpia] += 1

        return self._convertir_a_dfs(cat)
    
#Creamos los dataframes con los tipos de palabra y su contabilización.
#Hayamos sus frecuencias, porcentajes etc... Y acabamos ordenando
    def _convertir_a_dfs(self, categorias):
        tot_doc = sum(categorias["Total"].values())
        for c, counter in categorias.items():
            tot_cat = sum(counter.values())
            df = pd.DataFrame(counter.items(), columns=["Palabra", "Frecuencia"])
            df["Porcentaje en Categoría"] = (df["Frecuencia"] / tot_cat * 100) if tot_cat > 0 else 0
            if c != "Entidades (NER)":
                df["Porcentaje en Documento"] = (df["Frecuencia"] / tot_doc * 100) if tot_doc > 0 else 0
            categorias[c] = df.sort_values(by="Frecuencia", ascending=False)
        return categorias
#Basicamente analizamos cuando los contextos van relacionados a palabras positivas o negativas.
#Obtenemos el porcentaje de positivismo, negativismo y neutralidad (muy rudimentario pero le valió al chaval)
    def analizar_sentimiento(self, texto):
        texto_lower = texto.lower()
        pos = ['buen', 'mejor', 'excelente', 'positivo', 'crecimiento', 'derecho', 'paz', 'seguridad', 'acuerdo', 'favor', 'beneficio', 'avance', 'solución', 'apoyo', 'éxito']
        neg = ['mal', 'peor', 'negativo', 'crisis', 'engaño', 'insuficiente', 'problema', 'delito', 'corrupción', 'contra', 'riesgo', 'pérdida', 'crítica', 'rechazo', 'fracaso', 'violencia']
        c_pos = sum(len(re.findall(fr'\b{p}\b', texto_lower)) for p in pos)
        c_neg = sum(len(re.findall(fr'\b{n}\b', texto_lower)) for n in neg)
        c_neu = max((len(texto_lower.split()) // 15) - (c_pos + c_neg), 0)
        total = c_pos + c_neg + c_neu
        if total == 0: return {"Positivo": 33, "Negativo": 33, "Neutral": 34} 
        return {"Positivo": round((c_pos / total) * 100, 1), "Negativo": round((c_neg / total) * 100, 1), "Neutral": round((c_neu / total) * 100, 1)}

#Generamos la nube de palabras, las imágenes siempre van en la carpeta Static.
#Debemos convertir los dataframe en diccionarios ya que la librería de las nubes no capta los dataframes.
#Obtenemos la imagen y la guardamos en el lugar adecuado.
class GeneradorReportes:
    @staticmethod
    def generar_nube(df, nombre_imagen):
        try:
            os.makedirs('static', exist_ok=True)
            dicc = dict(zip(df['Palabra'], df['Frecuencia']))
            if not dicc: return None
            wc = WordCloud(width=800, height=400, background_color='white', colormap='viridis').generate_from_frequencies(dicc)
            plt.figure(figsize=(10, 5)); plt.imshow(wc, interpolation='bilinear'); plt.axis('off')
            plt.savefig(os.path.join('static', nombre_imagen)); plt.close()
            return nombre_imagen 
        except Exception as e: logging.error(f"Fallo en nube {nombre_imagen}: {e}"); return None

#Abrimos y escribimos el excel, Creamos cada una de las pestañas
#Agregamos los distintos dataframes de palabras en el excel.
#Hayamos también las columnas con los calculos de porcentajes y demás.
#Guardamos.
    @staticmethod
    def guardar_excel(dicc_dfs, nombre_archivo):
        try:
            with pd.ExcelWriter(nombre_archivo, engine='openpyxl') as w:
                for n, df in dicc_dfs.items(): df.to_excel(w, sheet_name=n, index=False)
                t_cat = {c: d["Frecuencia"].sum() for c, d in dicc_dfs.items() if c not in ["Total", "Entidades (NER)"]}
                df_res = pd.DataFrame(list(t_cat.items()), columns=["Categoría", "Total de Palabras"])
                t_gen = df_res["Total de Palabras"].sum()
                df_res["Porcentaje en Documento"] = (df_res["Total de Palabras"] / t_gen * 100) if t_gen > 0 else 0
                df_res.to_excel(w, sheet_name="Resumen Total", index=False)
        except Exception as e: logging.error(f"Fallo guardando Excel {nombre_archivo}: {e}")

#Esta es la función en la que aplicamos el diseño del excel
    @staticmethod
    def aplicar_formato_excel(archivo):
        try:
            wb = load_workbook(archivo)
            #Escala para los colores entre en verde y el rojo (modo semáforo)
            escala = ColorScaleRule(start_type='min', start_color='F8696B', mid_type='percentile', mid_value=50, mid_color='FFEB84', end_type='max', end_color='63BE7B')
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                ws.auto_filter.ref = ws.dimensions
                
                # Aplicación de este mismo formato para porcentajes y frecuencias
                for col in ws.columns:
                    c_hdr = ws.cell(row=1, column=col[0].column)
                    if c_hdr.value in ["Frecuencia", "Total de Palabras"] or "Porcentaje" in str(c_hdr.value) or "Ponderación" in str(c_hdr.value):
                        ws.conditional_formatting.add(f"{ws.cell(row=2, column=col[0].column).coordinate}:{ws.cell(row=ws.max_row, column=col[0].column).coordinate}", escala)
                    c_hdr.font = Font(bold=True); c_hdr.fill = PatternFill(start_color="D3D3D3", fill_type="solid")
                    ws.column_dimensions[col[0].column_letter].width = 22

                #Generación de los graficos de proporciones y top 10
                if ws.max_row > 1:
                    if sheet == "Resumen Total":
                        chart = PieChart()
                        chart.title = "Proporción de Categorías"
                        labels = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
                        data = Reference(ws, min_col=2, min_row=1, max_row=ws.max_row)
                        chart.add_data(data, titles_from_data=True)
                        chart.set_categories(labels)
                        ws.add_chart(chart, "E2")
                    elif sheet not in ["Final", "Total"]:
                        limite = min(11, ws.max_row)
                        chart = BarChart()
                        chart.title = f"Top 10 {sheet}"
                        chart.y_axis.title = 'Frecuencia'
                        labels = Reference(ws, min_col=1, min_row=2, max_row=limite)
                        data = Reference(ws, min_col=2, min_row=1, max_row=limite)
                        chart.add_data(data, titles_from_data=True)
                        chart.set_categories(labels)
                        ws.add_chart(chart, "F2")
            wb.save(archivo)
        except Exception as e: logging.error(f"Fallo formato Excel {archivo}: {e}")
#El constructor de orquestador se encarga de tener a mano todos los posibles procesos a realizar.
class OrquestadorApp:
    def __init__(self):
        self.archivos = GestorArchivos(); 
        self.nlp = ProcesadorNLP(); 
        self.reportes = GeneradorReportes()

    def procesar_corpus(self, r_fav, r_con, nom_fav, nom_con, nom_comb, uid):
        #Damos la extensión para excel al documento (se guarda con el nombre indicado, un identificador, y la extensión)
        n_fav = f"{nom_fav}_{uid}.xlsx" if not nom_fav.endswith('.xlsx') else f"{nom_fav.split('.xlsx')[0]}_{uid}.xlsx"
        n_con = f"{nom_con}_{uid}.xlsx" if not nom_con.endswith('.xlsx') else f"{nom_con.split('.xlsx')[0]}_{uid}.xlsx"
        n_total = f"{nom_comb}_{uid}.xlsx" if not nom_comb.endswith('.xlsx') else f"{nom_comb.split('.xlsx')[0]}_{uid}.xlsx"
        
        #Procedemos a lo mismo con la imagen del la nube de palabras, pero con formato .png
        img_fav_name = f"nube_fav_{uid}.png"; img_con_name = f"nube_con_{uid}.png"
        
        #Leemos los documentos, guardamos en texto plano, clasificamos las palabras y guardamos el excel
        t_f = self.archivos.leer_documento(r_fav); t_c = self.archivos.leer_documento(r_con)
        self.archivos.guardar_texto_plano(t_f.replace('\n', ' '), 'uploads/favor.txt')
        self.archivos.guardar_texto_plano(t_c.replace('\n', ' '), 'uploads/contra.txt')

        dfs_f = self.nlp.clasificar_palabras(t_f, self.nlp.limpiar_texto(t_f))
        dfs_c = self.nlp.clasificar_palabras(t_c, self.nlp.limpiar_texto(t_c))

        self.reportes.guardar_excel(dfs_f, n_fav)
        self.reportes.guardar_excel(dfs_c, n_con)
        
#Obtención de sumas, frecuencias y porcentajes. Guardamos el excel
        dfs_suma = {}
        t_gen = dfs_f["Total"]["Frecuencia"].sum() + dfs_c["Total"]["Frecuencia"].sum()
        for cat in ["Sustantivos", "Verbos", "Adjetivos", "Adverbios", "Entidades (NER)", "Total"]:
            df_sum = dfs_f[cat].set_index("Palabra")[["Frecuencia"]].add(dfs_c[cat].set_index("Palabra")[["Frecuencia"]], fill_value=0).reset_index()
            t_cat = df_sum["Frecuencia"].sum()
            df_sum["Porcentaje en Categoría"] = (df_sum["Frecuencia"] / t_cat * 100).fillna(0) if t_cat > 0 else 0
            if cat != "Entidades (NER)":
                df_sum["Porcentaje en Documento"] = (df_sum["Frecuencia"] / t_gen * 100).fillna(0) if t_gen > 0 else 0
            dfs_suma[cat] = df_sum.sort_values(by="Frecuencia", ascending=False)
            
        self.reportes.guardar_excel(dfs_suma, n_total)

        df_fav_f = dfs_f["Total"][["Palabra", "Porcentaje en Documento"]].rename(columns={"Porcentaje en Documento": "Porcentaje en Corpus_a_favor"})
        df_con_f = dfs_c["Total"][["Palabra", "Porcentaje en Documento"]].rename(columns={"Porcentaje en Documento": "Porcentaje en Corpus_en_contra"})
        df_fin = dfs_suma["Total"][["Palabra", "Porcentaje en Documento"]].merge(df_fav_f, on="Palabra", how="left").merge(df_con_f, on="Palabra", how="left").fillna(0)
        
        div = df_fin["Porcentaje en Corpus_a_favor"] + df_fin["Porcentaje en Corpus_en_contra"]
        df_fin["Ponderación Corpus_a_favor"] = (df_fin["Porcentaje en Corpus_a_favor"] / div * 100).fillna(0)
        df_fin["Ponderación Corpus_en_contra"] = (df_fin["Porcentaje en Corpus_en_contra"] / div * 100).fillna(0)
        
        with pd.ExcelWriter(n_total, mode="a", engine="openpyxl", if_sheet_exists="replace") as w:
            df_fin[df_fin["Porcentaje en Documento"] > 0.02].to_excel(w, sheet_name="Final", index=False)

        #Aplicamos el formato de excel
        for f in [n_fav, n_con, n_total]: self.reportes.aplicar_formato_excel(f)

        img_f = self.reportes.generar_nube(dfs_f["Total"], img_fav_name)
        img_c = self.reportes.generar_nube(dfs_c["Total"], img_con_name)

        #Creamos la barra de sentimientos
        sentimientos = {"favor": self.nlp.analizar_sentimiento(t_f), "contra": self.nlp.analizar_sentimiento(t_c)}
        #retornamos todos los datos
        return n_total, n_fav, n_con, img_f, img_c, sentimientos

    #Método que sirve para obtener los contextos de la palabra buscada, porcentajes, y total de menciones.
    def buscar_contexto(self, palabra, limite=5):
        res = {'favor': {'frases': [], 'conteo': 0, 'porcentaje': 0}, 'contra': {'frases': [], 'conteo': 0, 'porcentaje': 0}, 'total_menciones': 0}
        try:
            with open('uploads/favor.txt', 'r', encoding='utf-8') as f: txt_f = f.read()
            with open('uploads/contra.txt', 'r', encoding='utf-8') as f: txt_c = f.read()
            oc_f, oc_c = len(re.findall(f'(?i)({palabra})', txt_f)), len(re.findall(f'(?i)({palabra})', txt_c))
            res['favor']['conteo'] = oc_f; res['contra']['conteo'] = oc_c; res['total_menciones'] = oc_f + oc_c
            if res['total_menciones'] > 0:
                res['favor']['porcentaje'] = round((oc_f / res['total_menciones']) * 100, 1)
                res['contra']['porcentaje'] = round((oc_c / res['total_menciones']) * 100, 1)
            for key, txt in [('favor', txt_f), ('contra', txt_c)]:
                for f in [s.strip() for s in txt.split('.') if palabra.lower() in s.lower()][:limite]:
                    color = "text-blue-700 bg-blue-100" if key == 'favor' else "text-red-700 bg-red-100"
                    res[key]['frases'].append(re.sub(f'(?i)({palabra})', fr'<b class="{color} px-1 rounded">\1</b>', f) + ".")
        except: pass
        return res