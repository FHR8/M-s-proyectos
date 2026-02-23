#INSTALACIONES PREVIAS
#Instalación del entorno Visual Studio Code.
#Instalamos en VSC las extensiones PYTHON, Y PYLANCE (Mayor eficiencia)
#Python desde la página oficial, En mi caso 3.12. Aconsejable una versión actual pero nunca la última.
#Instalamos el pip, a partir del cual podremos instalar el resto de librerías python3 get-pip.py
#podemos verificar la versión que tenemos con pip --version
#Creamos el entorno para el proyecto python -m venv venv 
#Activamos el entorno .\venv\Scripts\activate
#Instalamos las librerias pandas, spacy, openpyxl, python-docx y pdfplumber necesarias, pip install pandas spacy openpyxl python-docx pdfplumber
#Instalamos el modelo español para Spacy. python -m spacy download es_core_news_sm


#Importamos la librería pandas para tratar textos.
import pandas as pd
#Importamos la librería colections para trabajar con contadores (Permite ir sumando las palabras)
from collections import Counter, defaultdict
#Importamos re para obtener expresiones regulares (como la del método limpiar texto)
import re
#importamos docx y pdfplumber que nos permite trabajar con los formatos word y pdf (pdfplumber especifico para trabajar con tablas)
from docx import Document
import pdfplumber
#Importamos os, que nos permite la interacción con el sistema operativo, para poder leer y crear documentos
import os
#Importamos spacy, que es una librería avanzada para trabajo con lenguaje natural. 
import spacy
#Importamos las stop words para a su debido momento eliminarlas del excel obtenido.
from spacy.lang.es.stop_words import STOP_WORDS
#Importamos openpyxl para darle formato al excel y crear los gráficos circulares.
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.chart import PieChart, Reference

# Cargamos el modelo de spaCy para español (spacy puede trabajar con multitud de lenguas)
nlp = spacy.load("es_core_news_sm")

# Diccionario de normalización para palabras específicas
normalizacion = {
    "derechos": "derecho",
    "leyes": "ley",
    "personas": "persona",
    "niños": "niño"
    # Agregar más términos según sea necesario
}

#obtenemos el texto completo del documento en formato word.
def leer_docx(archivo):
    doc = Document(archivo)
    texto = ''
    for parrafo in doc.paragraphs:
        texto += parrafo.text + '\n'
    return texto
#obtenemos el texto completo del documento en formato word.
def leer_pdf(archivo):
    texto = ''
    with pdfplumber.open(archivo) as pdf:
        for pagina in pdf.pages:
            texto += pagina.extract_text() + '\n'
    return texto

#Transformamos las palabras, todas a minúsculas para no diferenciar, elimina espacios extraños, y elimina todas las palabras que se encuentren en las stopwords*/
def limpiar_texto(texto):
    texto = texto.lower()
    texto = re.sub(r'[^a-záéíóúüñ\s]', '', texto)
    palabras = texto.split()
    palabras = [palabra for palabra in palabras if palabra not in STOP_WORDS]
    return " ".join(palabras)  
#Procesamos el texto con spacy 
def clasificar_palabras(texto):
    doc = nlp(texto)  
    #Obtenemos y contamos las palabras según las categorías.
    categorias = {"Sustantivos": Counter(), "Verbos": Counter(), "Adjetivos": Counter(), "Adverbios": Counter() , "Total": Counter()}
    #En caso de palabras ambiguas (puedan ser de varias categorías), analizamos el contexto y la clasificamos adecuadamente
    apariciones = defaultdict(lambda: Counter())  
    
    #Ignoramos signos de puntuación y caractéres extraños
    for token in doc:
        if not token.is_alpha: 
            continue
        lemma = token.lemma_
        
         # Aplicamos normalización si la palabra está en el diccionario
        if lemma in normalizacion:
            lemma = normalizacion[lemma]
        
        categoria = token.pos_
        
        #Todas las palabras nuevas, las insertamos en su pestaña y le damos el valor de 1
        if categoria == "NOUN":
            categorias["Sustantivos"][lemma] += 1
 # Evitar agregar pronombres personales como verbos
        elif categoria == "VERB":
             if " " in lemma:
                lemma = lemma.split(" ")[0]
             categorias["Verbos"][lemma] += 1
        elif categoria == "ADJ":
            categorias["Adjetivos"][lemma] += 1
        elif categoria == "ADV":
            categorias["Adverbios"][lemma] += 1
            
        categorias["Total"][lemma] += 1
        
        #En el caso de que esa palabra ya haya aparecido, aumentamos sus repeticiones en 1
        apariciones[lemma][categoria] += 1  
    #En la pestaña de todos los valores, realizamos lo anterior pero con todas las palabras.
    total_palabras_documento = sum(sum(counter.values()) for counter in categorias.values())
    
    # Calculamos el porcentaje de frecuencia dentro de cada categoría y del total del documento
    for categoria, counter in categorias.items():
        total_palabras_categoria = sum(counter.values())
        df = pd.DataFrame(counter.items(), columns=["Palabra", "Frecuencia"])
        df["Porcentaje en Categoría"] = (df["Frecuencia"] / total_palabras_categoria) * 100
        df["Porcentaje en Documento"] = (df["Frecuencia"] / total_palabras_documento) * 100
        df = df.sort_values(by="Frecuencia", ascending=False)
        categorias[categoria] = df
    
    return categorias

# procesamos el documento, dependiendo del formato, o recibimos el mensaje de error si no es un word o pdf
def procesar_documento(archivo):
    if archivo.endswith('.docx'):
        texto = leer_docx(archivo)
    elif archivo.endswith('.pdf'):
        texto = leer_pdf(archivo)
    else:
        raise ValueError("El archivo debe ser .docx o .pdf")
    #Limpiamos el texto (Eliminamos las stopwords, convertimos a minusculas...)
    texto_limpio = limpiar_texto(texto)
    #Clasificamos las palabras
    return clasificar_palabras(texto_limpio)

#
def aplicar_formato_excel(archivo):
    wb = load_workbook(archivo)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        
        # Aplicamos negrita a la primera fila,le damos el fondo gris, y lo centramos.
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            cell.fill = PatternFill(start_color="D3D3D3", fill_type="solid")
        
        # Ajustamos el tamaño de las columnas según la longitud de las palabras.
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = (max_length + 2)
            
            # Obtenemos el top de 10 palabras de cada tipo           
            data = Reference(ws, min_col=2, min_row=2, max_row=11)          
            labels = Reference(ws, min_col=1, min_row=2, max_row=11)
            #Creamos el diagrama circular.
            chart = PieChart()
            #Agregamos y damos tamaño a las 10 palabras más repetidas
            chart.add_data(data, titles_from_data=False)
            chart.set_categories(labels)
            chart.title = "Top 10 Palabras"
            # Ubicación del gráfico
            ws.add_chart(chart, "I2")  
            #Colocamos el gráfico.
            ws.column_dimensions[col_letter].width = adjusted_width
            #Guardamos el archivo.
    wb.save(archivo)

#Guardamos el excel con las pestañas de las 4 categorías (Escribiendolo palabra por palabra).
def guardar_excel_multiple(dfs, nombre_archivo):
    with pd.ExcelWriter(nombre_archivo) as writer:
        for categoria, df in dfs.items():
            df.to_excel(writer, sheet_name=categoria, index=False)
        
        # Se crea la hoja de resumen total de palabras por categoría
        total_por_categoria = {cat: sum(df["Frecuencia"]) for cat, df in dfs.items()}
        #Suma cada categoría
        total_general = sum(total_por_categoria.values())
        #Damos el formato a la pestaña
        df_totales = pd.DataFrame(list(total_por_categoria.items()), columns=["Categoría", "Total de Palabras"])
        df_totales["Porcentaje en Documento"] = (df_totales["Total de Palabras"] / total_general) * 100
        df_totales.to_excel(writer, sheet_name="Resumen Total", index=False)
    #Mensaje de que se ha realizado correctamente el guardado
        
    print(f"Se ha guardado: {nombre_archivo}")

#PARTE MODIFICABLE POR EL USUARIO, indicamos los 2 archivos que vamos a tratar. 
archivo1 = "Corpus a favor.docx"
archivo2 = "Corpus en contra.docx"

#Los procesamos con los pasos vistos en los métodos anteriores.
dfs1 = procesar_documento(archivo1)
dfs2 = procesar_documento(archivo2)

#Lo guardamos todo en el excel.
guardar_excel_multiple(dfs1, "Corpus_a_favor.xlsx")
guardar_excel_multiple(dfs2, "Corpus_en_contra.xlsx")

#Le damos el formato al excel
aplicar_formato_excel("Corpus_a_favor.xlsx")
aplicar_formato_excel("Corpus_en_contra.xlsx")

# Sumamos frecuencias de ambos documentos
dfs_suma = {}
total_palabras_documento = sum(sum(df["Frecuencia"]) for df in dfs1.values()) + sum(sum(df["Frecuencia"]) for df in dfs2.values())
for categoria in ["Sustantivos", "Verbos", "Adjetivos", "Adverbios", "Total"]:
    df1 = dfs1[categoria].set_index("Palabra")
    df2 = dfs2[categoria].set_index("Palabra")
    df_suma = df1.add(df2, fill_value=0).reset_index()
    total_palabras_categoria = df_suma["Frecuencia"].sum()
    df_suma["Porcentaje en Categoría"] = (df_suma["Frecuencia"] / total_palabras_categoria) * 100
    df_suma["Porcentaje en Documento"] = (df_suma["Frecuencia"] / total_palabras_documento) * 100
    df_suma = df_suma.sort_values(by="Frecuencia", ascending=False)
    dfs_suma[categoria] = df_suma
 
#Guardamos el tercer documento, que combina los 2 anteriores.
guardar_excel_multiple(dfs_suma, "Frecuencia_total.xlsx")

# Cargar la pestaña "Total" de Corpus_a_favor.xlsx
df_favor = pd.read_excel("Corpus_a_favor.xlsx", sheet_name="Total")[["Palabra", "Porcentaje en Documento"]]
df_favor = df_favor.rename(columns={"Porcentaje en Documento": "Porcentaje en Corpus_a_favor"})

# Cargar la pestaña "Total" de Corpus_en_contra.xlsx
df_contra = pd.read_excel("Corpus_en_contra.xlsx", sheet_name="Total")[["Palabra", "Porcentaje en Documento"]]
df_contra = df_contra.rename(columns={"Porcentaje en Documento": "Porcentaje en Corpus_en_contra"})

# Cargar la pestaña "Total" de Frecuencia_total.xlsx y seleccionar columnas necesarias
df_total = dfs_suma["Total"][["Palabra", "Porcentaje en Documento"]]

# Unir los tres DataFrames en base a la palabra
df_final = df_total.merge(df_favor, on="Palabra", how="left")
df_final = df_final.merge(df_contra, on="Palabra", how="left")

# Calcular las ponderaciones
df_final["Ponderación Corpus_a_favor"] = (df_final["Porcentaje en Corpus_a_favor"] / (df_final["Porcentaje en Corpus_a_favor"] + df_final["Porcentaje en Corpus_en_contra"])) * 100
df_final["Ponderación Corpus_en_contra"] = (df_final["Porcentaje en Corpus_en_contra"] / (df_final["Porcentaje en Corpus_a_favor"] + df_final["Porcentaje en Corpus_en_contra"])) * 100

# Filtrar palabras con porcentaje en documento mayor a 0.02
threshold = 0.02
df_final = df_final[df_final["Porcentaje en Documento"] > threshold]

# Formatear las columnas numéricas como porcentaje
#df_final.iloc[:, 1:] = df_final.iloc[:, 1:].applymap(lambda x: "{:.2f}%".format(float(str(x).replace(',', '.'))))

# Guardamos en la pestaña "Final"
with pd.ExcelWriter("Frecuencia_total.xlsx", mode="a", if_sheet_exists="replace") as writer:
    df_final.to_excel(writer, sheet_name="Final", index=False)
    
#Y le damos el formato de excel.
aplicar_formato_excel("Frecuencia_total.xlsx")
